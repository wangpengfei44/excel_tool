import json
from collections.abc import Generator
from typing import Any

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from io import BytesIO


class ArrayToExcelTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        """
        Convert a JSON-encoded 2D array into an Excel file (.xlsx),
        with optional support for column widths, merged cells,
        cell styles (including region-level configuration), and row heights.

        Parameters:
        - tool_parameters: dictionary containing a field "data_json",
          which is a JSON string with the following fields:
            * data: 2D array representing the Excel table content
            * col_widths: dict mapping column numbers (as strings) to widths (optional)
            * merges: list of merged cell ranges, each item with
                start_row, start_col, end_row, end_col
            * cell_styles: list of cell style definitions, each item includes:
                - start_row, start_col, end_row (optional), end_col (optional)
                - style: dict with the following optional keys:
                  - alignment: horizontal alignment ("center" / "left" / "right")
                  - vertical: vertical alignment ("center" / "top" / "bottom") (default: "center")
                  - font_size: integer font size
                  - bold: boolean indicating bold text
                  - bgcolor: hex RGB background color (e.g., "FFFF00")
                  - border: boolean indicating whether to apply thin border
                  - wrap_text: whether to enable text wrapping
            * row_heights: dict mapping row numbers (as strings) to row heights (optional)

        Returns:
        - Generator yielding ToolInvokeMessage.
          On success, returns binary Excel file.
          On failure, returns a text message.
        """

        data_json = tool_parameters.get("data_json", "")
        if not data_json:
            yield self.create_text_message("Invalid parameters: 'data_json' is missing")
            return

        try:
            if isinstance(data_json, str):
                params = json.loads(data_json)
                if isinstance(params, str):
                    params = json.loads(params)
            else:
                params = data_json
        except Exception as e:
            yield self.create_text_message(f"Failed to parse parameters: {e}")
            return

        data = params.get("data")
        col_widths = params.get("col_widths", {})
        merges = params.get("merges", [])
        cell_styles = params.get("cell_styles", [])  # Ensure it's a list
        row_heights = params.get("row_heights", {})

        if not data or not isinstance(data, list):
            yield self.create_text_message("Invalid parameters: 'data' should be a 2D array")
            return

        wb = Workbook()
        ws = wb.active

        # Define thin border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for r, row in enumerate(data, start=1):
            if not isinstance(row, list):
                yield self.create_text_message(f"Invalid parameters: row {r} is not a list")
                return
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(vertical='center', wrap_text=True)  # Default: vertical center & wrap text
                cell.border = thin_border  # Default thin border

        # Apply cell styles; each style block includes range and style dict
        for style_range in cell_styles:
            try:
                start_row = style_range["start_row"]
                start_col = style_range["start_col"]
                end_row = style_range.get("end_row", start_row)
                end_col = style_range.get("end_col", start_col)
                style_cfg = style_range["style"]
            except Exception as e:
                yield self.create_text_message(f"Style config error: {e}")
                return

            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    cell = ws.cell(row=r, column=c)

                    # Set alignment; default vertical=center; wrap_text default True
                    align_val = style_cfg.get("alignment")
                    vertical_val = style_cfg.get("vertical", "center")
                    wrap_text_val = style_cfg.get("wrap_text", True)
                    if align_val:
                        cell.alignment = Alignment(horizontal=align_val, vertical=vertical_val,
                                                   wrap_text=wrap_text_val)
                    else:
                        cell.alignment = Alignment(vertical=vertical_val, wrap_text=wrap_text_val)

                    # Set font
                    font_args = {}
                    if "font_size" in style_cfg:
                        font_args["size"] = style_cfg["font_size"]
                    if style_cfg.get("bold"):
                        font_args["bold"] = True
                    if font_args:
                        cell.font = Font(**font_args)

                    # Set background color
                    bgcolor = style_cfg.get("bgcolor")
                    if bgcolor:
                        cell.fill = PatternFill(fill_type="solid", fgColor=bgcolor)

                    # Control borders: if "border" is False, remove border
                    if "border" in style_cfg:
                        if style_cfg.get("border"):
                            cell.border = thin_border
                        else:
                            cell.border = Border()  # No border

        # Set row heights; keys are strings and need to be converted to int
        try:
            for row_str, height in row_heights.items():
                row_idx = int(row_str)
                ws.row_dimensions[row_idx].height = float(height)
        except Exception as e:
            yield self.create_text_message(f"Failed to set row height: {e}")
            return

        # Set column widths for columns A-Z (1-26)
        try:
            for col_str, width in col_widths.items():
                col_idx = int(col_str)
                if 1 <= col_idx <= 26:
                    col_letter = chr(64 + col_idx)
                    ws.column_dimensions[col_letter].width = float(width)
                else:
                    yield self.create_text_message(f"Unsupported column index (max 26): {col_idx}")
                    return
        except Exception as e:
            yield self.create_text_message(f"Failed to set column width: {e}")
            return

        # Apply merged cells
        try:
            for merge in merges:
                start_row = merge["start_row"]
                start_col = merge["start_col"]
                end_row = merge["end_row"]
                end_col = merge["end_col"]
                start_cell = ws.cell(row=start_row, column=start_col).coordinate
                end_cell = ws.cell(row=end_row, column=end_col).coordinate
                ws.merge_cells(f"{start_cell}:{end_cell}")
        except Exception as e:
            yield self.create_text_message(f"Failed to merge cells: {e}")
            return

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        yield self.create_blob_message(
            blob=output.getvalue(),
            meta={
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "output_filename": "output.xlsx",
            },
        )
